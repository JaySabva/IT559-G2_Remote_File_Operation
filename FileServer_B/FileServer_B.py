import os
import xmlrpc.client
from xmlrpc.server import SimpleXMLRPCServer
import openpyxl
import threading
import heapq

hostID = "192.168.6.6"
FileServer_B = SimpleXMLRPCServer((hostID, 9002), logRequests=True, allow_none=True)

class MinHeap:
    def __init__(self):
        self.heap = []

    def push(self, item):
        heapq.heappush(self.heap, item)

    def pop(self):
        return heapq.heappop(self.heap)

    def peek(self):
        return self.heap[0] if self.heap else None

    def empty(self):
        return len(self.heap) == 0

server_heap = {}
server_file = "Servers.xlsx"  # Assuming Servers.xlsx contains backup server details
if os.path.exists(server_file):
    server_workbook = openpyxl.load_workbook(server_file)
    server_worksheet = server_workbook.active
else:
    print("Server file not found!")

def write(filename, data, primary, flag, timestamp=None):
    mode = 'a' if (os.path.exists(filename) and flag) else 'w'
    with open(filename, mode) as file:
        file.write(data + "\n")
    if primary:
        backup_thread = threading.Thread(target=send_to_backups, args=(filename, data, mode, timestamp))
        backup_thread.start()
        return True
    return True

def send_to_backups(filename, data, mode, timestamp):
    backup_servers = []
    print("Sending data to backup servers")
    if os.path.exists(server_file):
        server_rows = list(server_worksheet.iter_rows(values_only=True))
        for row in server_rows[1:]:
            if row[2] != 9002:  # Exclude FileServer B's port from backups
                addr = hostID
                port = row[2]
                if (addr, port) not in server_heap:
                    server_heap[(addr, port)] = MinHeap()
                server_heap[(addr, port)].push((timestamp, filename, data, mode))
                try:
                    proxy = xmlrpc.client.ServerProxy(f"http://{addr}:{port}/", allow_none=True)
                    while not server_heap[(addr, port)].empty():
                        timestamp, filename, data, mode = server_heap[(addr, port)].peek()
                        response = proxy.write(filename, data, False, mode == 'a', timestamp)
                        if response:
                            server_heap[(addr, port)].pop()
                            backup_servers.append((filename, addr, port, timestamp))
                except Exception as e:
                    print(f"Error connecting to {addr}:{port}: {e}")
                    # Handle connection errors here if needed
    else:
        return False

    # Send information about backup servers to the master server
    master_proxy = xmlrpc.client.ServerProxy(f"http://{hostID}:9000/", allow_none=True)
    master_proxy.send_backup_servers(backup_servers)

    return True

def read(filename):
    if os.path.exists(filename):
        with open(filename, 'r') as file:
            return file.read()
    else:
        return False

def getMyUpdate():
    print("getMyUpdate called")
    server_rows = list(server_worksheet.iter_rows(values_only=True))
    for row in server_rows[1:]:
        addr = hostID
        port = row[2]
        if port != 9002:
            try:
                proxy = xmlrpc.client.ServerProxy(f"http://{addr}:{port}/", allow_none=True)
                response = proxy.sendUpdate(addr, 9002)
                for update in response:
                    filename, data, mode, timestamp = update
                    print(filename, data, mode, timestamp)
                    write(filename, data, False, mode == 'a', timestamp)
                print("getMyUpdate done")
            except Exception as e:
                print(f"Error connecting to {addr}:{port}: {e}")

def sendUpdate(addr, port):
    print(f"Sending update to {addr}:{port}")
    updates = []
    if (addr, port) in server_heap and not server_heap[(addr, port)].empty():
        while not server_heap[(addr, port)].empty():
            timestamp, filename, data, mode = server_heap[(addr, port)].pop()
            updates.append((filename, data, mode, timestamp))

    return updates

def initialize_getMyUpdate():
    global getMyUpdate_executed
    if not getMyUpdate_executed:
        getMyUpdate()
        getMyUpdate_executed = True

getMyUpdate_executed = False
initialize_getMyUpdate()

FileServer_B.register_function(write, "write")
FileServer_B.register_function(read, "read")
FileServer_B.register_function(getMyUpdate, "getMyUpdate")
FileServer_B.register_function(sendUpdate, "sendUpdate")

print("FileServer B running on localhost:9002")
FileServer_B.serve_forever()
